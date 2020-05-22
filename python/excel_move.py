import openpyxl
import pprint

if __name__ == '__main__':

    book = openpyxl.load_workbook('example.xlsx')
    sheet_names = book.sheetnames

    # print(sheet_names)

    column_date  = 5

    sheet_name1 = 'テスト　太郎'

    sheet1 = book[sheet_name1]
    sheet2 = book[sheet_names[1]]
    text = """21:29 2020/05/22
テスト　太郎
項目1
0件
1件
25件
項目2
12件
55件
45件"""

    pprint.pprint(list(sheet1.values))

    input_text = text.splitlines()
    # print(input_text)

    for i, row in enumerate(input_text,1):
        print(i,row)
        sheet2.cell( i, column_date).value = row

    book.save('example.xlsx')