import openpyxl
import pandas as pd


if __name__ == '__main__':
    """
    Excelファイル(xlsx)を作るだけ
    `example.xlsx` が実行したカレントにできる
    """
    # wb = openpyxl.Workbook()

    # 保存
    # wb.save('example.xlsx')

    wb = openpyxl.load_workbook('example.xlsx')
    sheet_name = 'テスト　太郎'
    ws = wb[sheet_name]

    print(ws)

    ws_clums = 5


    ws.cell(2,ws_clums).value = 100

    wb.save('example.xlsx')

    wb.close()